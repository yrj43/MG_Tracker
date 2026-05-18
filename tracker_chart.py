import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import datetime, timedelta
from fuzzywuzzy import process

# ══════════════════════════════════════════════════════════════════════════════
#  LOOKUP LISTS  (edit these to match your actual data)
# ══════════════════════════════════════════════════════════════════════════════

DEFAULT_NAMES = [
    "Abarajithan Govindarajan",
    "Aishwarya Rajamohan",
    "Akash P M",
    "Akash.M Murali",
    "Akshay Kumar P",
    "Akshay V Kumar",
    "Allan Augustine",
    "Anshuman Dey",
    "Anusree Anil",
    "Archana Venkatesan",
    "Ashwin Kumar",
    "Avi Sharma",
    "Bala Thirupathi Raaja",
    "Barathi Priya",
    "Bhavani Dhanabalan",
    "Deepika Raghuraj",
    "Deepika Subramani",
    "Devanathan",
    "Dilip Suresh",
    "Divya Dharshini",
    "Divya Shree",
    "Durairaj Saravanakumar",
    "GN Karthik",
    "Gnana Jenifer Wilciya",
    "Gurumoorthy Vijayarangan",
    "Hemavathy R",
    "V Jahnavi",
    "Jeff Rohit",
    "Joshni N",
    "Karthick Gurunathan",
    "Karthik A",
    "Karthikeyan Panachavaranam",
    "Keerthana B",
    "Kiranraj Ravichandran",
    "Kishore S",
    "Kumaran Ramachandran",
    "Maha S",
    "Mariya Antony Britto",
    "Md Shadman Hayat Siddhiquie",
    "Meenakshi Maragathavel",
    "Melwin Manoj",
    "Mohammed Wihaj",
    "Moneshwar Devaraj",
    "Mukul Vyas Parameswar",
    "Nadhiya Siva Subramanian",
    "Najir Hussain Nashim Miyan",
    "Neelufur Begam",
    "Nishanthini Umapathy",
    "Palani Raja Vellaisamy",
    "Parthasarathy Letchumanan",
    "Prabhakaran Sekar",
    "Pragadeeshwaran Ganesan",
    "Priya Dharshini K",
    "Priyea Dharshani B",
    "Ritesh Suresh",
    "Rohit Subramani",
    "Rojini.S Sathish Kumar",
    "Sabariraj Iyyappan",
    "Sachin Rajesh",
    "Samyuktha Balakrishnaian",
    "Saquib Tanweer",
    "Sarathirajan K",
    "SarathKumar Ravikumar",
    "Sathish Kumar Venkatesan",
    "Shalini S",
    "Shalini Subramanian",
    "Shantha Kumar Saravanan",
    "Sivasankari Arumugam",
    "Sonia Selva Kumar",
    "Sridevi Rangarajan",
    "Sruthi Mathivanan",
    "Steffin T M",
    "Tamilarasi Balamurugan",
    "Veera Sabarinathan",
    "Vignesh Murugan",
    "Vijay Kumar R",
    "Vijayalakshmi Dhanabalan",
    "Vijayalakshmi Janakiraman",
    "Vinod Ram",
    "Vishwa Alagiri",
    "Yazhini Krishnamoorthy",
    "Yuvaraj Selvam"
]

DEFAULT_REGIONS = [
    "Atlantic", "Northeast", "Midwest",
    "Southeast", "Central","Pacific",
]

DEFAULT_ORDER_TYPES = [
    "Full Contract", "Ad Hoc", "BMG" , "GRP" , "One Connect" , "Political",
    
]

DEFAULT_ORDER_STATUSES = [
    "No Action Taken", "TF Action", "Completed", "CM Action",
]

EXPECTED_COLUMNS = [
    "Date", "Month", "Name", "Region", "TIM#", "CM", "Parameters",
    "MG Lines", "Oversold Lines", "Total MGL $", "Order Type",
    "Order Staus", "Reason", "Quality Met",
    "Error Type", "Error Comments", "TAT Miss", "Feedback Delivered by", "Year",
]

FUZZY_THRESHOLD = 80   # match score 0-100; below this = no match → red

# ══════════════════════════════════════════════════════════════════════════════
#  ORDER TYPE  — explicit keyword → canonical mapping (case-insensitive)
#  Checked BEFORE fuzzy matching so these always take priority.
# ══════════════════════════════════════════════════════════════════════════════

ORDER_TYPE_KEYWORD_MAP = {
    # ── GRP ──────────────────────────────────────────────────────────────────
    "ratings":          "GRP",
    "rating":           "GRP",
    "bmg in ratings":   "GRP",
    "grb":              "GRP",
    "grp":              "GRP",
    "rate":             "GRP",
    # ── Political ─────────────────────────────────────────────────────────────
    "fc - pol":         "Political",
    "ad hoc - pol":     "Political",
    "pol":              "Political",
    "political":        "Political",
    "bmg in pol":       "Political",
    # ── BMG ───────────────────────────────────────────────────────────────────
    "bmg in ad hoc":        "BMG",
    "bmg in fc":            "BMG",
    "bmg":                  "BMG",
    "bmg in full contract": "BMG",
    "bmg in fullcontract":  "BMG",
    # ── One Connect ───────────────────────────────────────────────────────────
    "oc":           "One Connect",
    "oneconnect":   "One Connect",
    "one-connect":  "One Connect",
    "one connect":  "One Connect",
}


# ══════════════════════════════════════════════════════════════════════════════
#  HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def is_empty(val):
    return pd.isna(val) or str(val).strip() == ""


def is_numeric_string(val):
    """True if val looks like a plain number (after stripping spaces)."""
    try:
        float(str(val).replace(",", "").replace("$", "").strip())
        return True
    except ValueError:
        return False


def fuzzy_match(value, valid_list, threshold=FUZZY_THRESHOLD):
    """
    Returns (matched_value, is_good_match).
    is_good_match=False means score < threshold → caller should flag red.
    """
    if is_empty(value):
        return None, False
    val_str = str(value).strip()
    if not valid_list:
        return val_str, False
    match, score = process.extractOne(val_str, valid_list)
    if score >= threshold:
        return match, True
    return val_str, False   # keep original but signal red


def get_date_info():
    """Return (target_date_str, month_str, year_int) based on today."""
    today = datetime.today()
    target = today - timedelta(days=1)
    date_str  = target.strftime("%d-%b-%Y")
    year      = target.year

    # Month rule: if today is the 1st, use previous month
    if today.day == 1:
        first_of_month = today.replace(day=1)
        prev_month     = first_of_month - timedelta(days=1)
        month_str      = prev_month.strftime("%B")
    else:
        month_str = target.strftime("%B")

    return date_str, month_str, year


# ══════════════════════════════════════════════════════════════════════════════
#  CORE CLEANING
# ══════════════════════════════════════════════════════════════════════════════

def clean_tracker(raw_df: pd.DataFrame,
                  valid_names: list,
                  valid_regions: list,
                  valid_order_types: list,
                  valid_order_statuses: list,
                  threshold: int = FUZZY_THRESHOLD):
    """
    Returns:
        df          – cleaned DataFrame
        red_cells   – set of (row_index, col_name) to highlight red
    """
    date_str, month_str, current_year = get_date_info()

    # ── Normalise incoming column names ──────────────────────────────────────
    raw_df.columns = [str(c).strip() for c in raw_df.columns]

    # ── Handle merged cells: openpyxl fills merged cells with NaN for the
    #    non-anchor cells. We forward-fill to capture the anchor value, then
    #    we only keep the FIRST occurrence per merged block.
    #    For the Name column specifically the spec says keep only the cell the
    #    person actually typed; duplicates from merging become "-".
    #    We achieve this by NOT forward-filling Name — blanks stay blank and
    #    get "-" later. All other cols are fine as-is from pandas read_excel.

    # ── Ensure every expected column exists ──────────────────────────────────
    for col in EXPECTED_COLUMNS:
        if col not in raw_df.columns:
            raw_df[col] = None

    df = raw_df[EXPECTED_COLUMNS].copy()

    red_cells = set()   # (row_index, col_name)

    # ── 1. DATE  ─────────────────────────────────────────────────────────────
    #    Always overwrite with (current date - 1) in required format
    df["Date"] = date_str

    # ── 2. MONTH ─────────────────────────────────────────────────────────────
    #    Always overwrite with computed month (handles 1st-of-month rule)
    df["Month"] = month_str

    # ── 3. NAME ──────────────────────────────────────────────────────────────
    def clean_name(val):
        if is_empty(val):
            return ("-", False)           # empty → "-", no red
        matched, good = fuzzy_match(val, valid_names, threshold)
        if good:
            return (matched, False)       # corrected name
        return (str(val).strip(), True)   # unrecognised → red

    name_results = df["Name"].apply(clean_name)
    df["Name"]   = name_results.apply(lambda x: x[0])
    for i, (_, red) in enumerate(name_results):
        if red:
            red_cells.add((i, "Name"))

    # ── 4. REGION ────────────────────────────────────────────────────────────
    def clean_region(val):
        if is_empty(val):
            return ("-", False)
        matched, good = fuzzy_match(val, valid_regions, threshold)
        if good:
            return (matched, False)
        return (str(val).strip(), True)

    region_results = df["Region"].apply(clean_region)
    df["Region"]   = region_results.apply(lambda x: x[0])
    for i, (_, red) in enumerate(region_results):
        if red:
            red_cells.add((i, "Region"))

    # ── 5. TIM# ──────────────────────────────────────────────────────────────
    #    Remove spaces; strip .0 from floats (e.g. 4494128.0 -> 4494128)
    def clean_tim(val):
        if is_empty(val):
            return "-"
        s = str(val).strip()
        try:
            f = float(s.replace(" ", ""))
            if f == int(f):
                return str(int(f))
        except ValueError:
            pass
        return s.replace(" ", "")

    df["TIM#"] = df["TIM#"].apply(clean_tim)

    # ── 6. CM ────────────────────────────────────────────────────────────────
    #    Leave as-is; empty → "-"
    df["CM"] = df["CM"].apply(lambda v: "-" if is_empty(v) else str(v).strip())

    # ── 7. PARAMETERS ────────────────────────────────────────────────────────
    df["Parameters"] = df["Parameters"].apply(
        lambda v: "-" if is_empty(v) else str(v).strip()
    )

    # ── 8. MG LINES / OVERSOLD LINES ─────────────────────────────────────────
    def clean_lines(val):
        if is_empty(val):
            return 0
        try:
            return int(float(str(val).replace(",", "").strip()))
        except ValueError:
            return 0   # unwanted text → 0

    df["MG Lines"]       = df["MG Lines"].apply(clean_lines)
    df["Oversold Lines"] = df["Oversold Lines"].apply(clean_lines)

    # ── 9. TOTAL MGL $ ───────────────────────────────────────────────────────
    def clean_currency(val):
        if is_empty(val):
            return 0.0
        try:
            return float(str(val).replace(",", "").replace("$", "").strip())
        except ValueError:
            return 0.0   # unwanted text → $0.00

    df["Total MGL $"] = df["Total MGL $"].apply(clean_currency)

    # ── 10. ORDER TYPE ───────────────────────────────────────────────────────
    # Priority 1: explicit keyword map (case-insensitive, stripped)
    # Priority 2: fuzzy match against valid list
    # Priority 3: unrecognised → flag red
    def clean_order_type(val):
        if is_empty(val):
            return ("-", False)
        val_stripped = str(val).strip()
        val_lower    = val_stripped.lower()

        # Check keyword map first
        if val_lower in ORDER_TYPE_KEYWORD_MAP:
            return (ORDER_TYPE_KEYWORD_MAP[val_lower], False)

        # Fall back to fuzzy match
        matched, good = fuzzy_match(val_stripped, valid_order_types, threshold)
        if good:
            return (matched, False)

        return (val_stripped, True)   # unrecognised → red

    ot_results      = df["Order Type"].apply(clean_order_type)
    df["Order Type"] = ot_results.apply(lambda x: x[0])
    for i, (_, red) in enumerate(ot_results):
        if red:
            red_cells.add((i, "Order Type"))

    # ── 11. ORDER STATUS ─────────────────────────────────────────────────────
    def clean_status(val):
        if is_empty(val):
            return ("-", False)
        matched, good = fuzzy_match(val, valid_order_statuses, threshold)
        if good:
            return (matched, False)
        return (str(val).strip(), True)

    os_results        = df["Order Staus"].apply(clean_status)
    df["Order Staus"] = os_results.apply(lambda x: x[0])
    for i, (_, red) in enumerate(os_results):
        if red:
            red_cells.add((i, "Order Staus"))

    # ── 13. REASON ───────────────────────────────────────────────────────────
    df["Reason"] = df["Reason"].apply(
        lambda v: "-" if is_empty(v) else str(v).strip()
    )

    # ── 14. QUALITY MET ──────────────────────────────────────────────────────
    df["Quality Met"] = "Yes"

    # ── 15. ERROR TYPE / ERROR COMMENTS ──────────────────────────────────────
    df["Error Type"]     = "-"
    df["Error Comments"] = "-"

    # ── 16. TAT MISS ─────────────────────────────────────────────────────────
    df["TAT Miss"] = "No"

    # ── 17. FEEDBACK DELIVERED BY ─────────────────────────────────────────────
    df["Feedback Delivered by"] = "-"

    # ── 18. YEAR ─────────────────────────────────────────────────────────────
    df["Year"] = current_year

    # ── 19. DROP rows where Name is "-" ─────────────────────────────────────
    #    After all cleaning, rows with Name="-" were originally empty → remove them.
    #    Also update red_cells row indices to match the new df after dropping.
    keep_mask = df["Name"] != "-"
    dropped_indices = set(df.index[~keep_mask].tolist())
    df = df[keep_mask].reset_index(drop=True)

    # Remap red_cell row indices: remove dropped rows, shift remaining rows up
    new_red_cells = set()
    for (row_i, col) in red_cells:
        if row_i not in dropped_indices:
            new_row = row_i - sum(1 for d in dropped_indices if d < row_i)
            new_red_cells.add((new_row, col))
    red_cells = new_red_cells

    return df, red_cells


# ══════════════════════════════════════════════════════════════════════════════
#  PIVOT TABLE
# ══════════════════════════════════════════════════════════════════════════════

def build_pivot(df: pd.DataFrame) -> pd.DataFrame:
    """Name | Count of TIM# | Sum of MG Lines | Sum of Oversold Lines"""
    pivot = (
        df.groupby("Name", as_index=False)
        .agg(
            **{
                "Count of TIM#":         ("TIM#",          "count"),
                "Sum of MG Lines":       ("MG Lines",       "sum"),
                "Sum of Oversold Lines": ("Oversold Lines", "sum"),
            }
        )
        .sort_values("Name")
        .reset_index(drop=True)
    )
    # Grand Total row
    grand = pd.DataFrame([{
        "Name":                  "Grand Total",
        "Count of TIM#":         pivot["Count of TIM#"].sum(),
        "Sum of MG Lines":       pivot["Sum of MG Lines"].sum(),
        "Sum of Oversold Lines": pivot["Sum of Oversold Lines"].sum(),
    }])
    pivot = pd.concat([pivot, grand], ignore_index=True)
    return pivot


# ══════════════════════════════════════════════════════════════════════════════
#  CHART DATA HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def build_region_data(df: pd.DataFrame) -> pd.DataFrame:
    """Region | Total Lines (MG + Oversold combined) | TIM Count"""
    region_df = df[df["Region"] != "-"].copy()
    grouped = (
        region_df.groupby("Region", as_index=False)
        .agg(
            **{
                "MG Lines Sum":       ("MG Lines",      "sum"),
                "Oversold Lines Sum": ("Oversold Lines", "sum"),
                "TIM Count":         ("TIM#",           "count"),
            }
        )
        .sort_values("Region")
        .reset_index(drop=True)
    )
    grouped["Total Lines"] = grouped["MG Lines Sum"] + grouped["Oversold Lines Sum"]
    grouped = grouped[["Region", "Total Lines", "TIM Count"]]
    return grouped


def build_order_type_data(df: pd.DataFrame) -> pd.DataFrame:
    """Order Type | Count"""
    ot_df = df[df["Order Type"] != "-"].copy()
    grouped = (
        ot_df.groupby("Order Type", as_index=False)
        .agg(**{"Count": ("Order Type", "count")})
        .sort_values("Count", ascending=False)
        .reset_index(drop=True)
    )
    return grouped


def build_order_status_data(df: pd.DataFrame) -> pd.DataFrame:
    """Order Status | Count"""
    os_df = df[df["Order Staus"] != "-"].copy()
    grouped = (
        os_df.groupby("Order Staus", as_index=False)
        .agg(**{"Count": ("Order Staus", "count")})
        .sort_values("Count", ascending=False)
        .reset_index(drop=True)
    )
    return grouped


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL EXPORT  — 3 sheets: cleaned data + pivot + charts
# ══════════════════════════════════════════════════════════════════════════════

def to_excel_bytes(df: pd.DataFrame, red_cells: set,
                   pivot: pd.DataFrame, tracker_sheet_name: str) -> bytes:
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.series import DataPoint

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:

        # ── Sheet 1 : Cleaned Tracker (all original styling preserved) ───────
        df.to_excel(writer, index=False, sheet_name=tracker_sheet_name)
        ws = writer.sheets[tracker_sheet_name]

        # styles
        hdr_fill  = PatternFill("solid", fgColor="1F3864")
        hdr_font  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        thin      = Side(style="thin", color="C8D0E0")
        border    = Border(left=thin, right=thin, top=thin, bottom=thin)
        alt_fill  = PatternFill("solid", fgColor="EEF2FF")
        wht_fill  = PatternFill("solid", fgColor="FFFFFF")
        red_fill  = PatternFill("solid", fgColor="FF3B30")

        currency_idx = EXPECTED_COLUMNS.index("Total MGL $") + 1  # 1-based

        # header
        for cell in ws[1]:
            cell.fill      = hdr_fill
            cell.font      = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)
            cell.border    = border

        # data rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            data_row = row_idx - 2          # 0-based
            base_fill = alt_fill if row_idx % 2 == 0 else wht_fill

            for cell in row:
                col_name = EXPECTED_COLUMNS[cell.column - 1]
                is_red   = (data_row, col_name) in red_cells

                cell.fill      = red_fill if is_red else base_fill
                cell.font      = Font(name="Calibri", size=10,
                                      bold=is_red,
                                      color="FFFFFF" if is_red else "000000")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border    = border

                if cell.column == currency_idx:
                    cell.number_format = '"$"#,##0.00'

        # column widths
        for col_cells in ws.columns:
            max_len = max(
                (len(str(c.value)) if c.value is not None else 0 for c in col_cells),
                default=10,
            )
            ws.column_dimensions[
                get_column_letter(col_cells[0].column)
            ].width = min(max_len + 4, 32)

        ws.row_dimensions[1].height = 34
        ws.freeze_panes = "A2"

        # ── Sheet 2 : Pivot ───────────────────────────────────────────────────
        pivot.to_excel(writer, index=False, sheet_name="Pivot")
        wp = writer.sheets["Pivot"]

        piv_hdr_fill = PatternFill("solid", fgColor="1F3864")
        piv_hdr_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        total_fill   = PatternFill("solid", fgColor="2E4057")
        total_font   = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        piv_alt      = PatternFill("solid", fgColor="EEF2FF")
        piv_wht      = PatternFill("solid", fgColor="FFFFFF")

        # pivot header
        for cell in wp[1]:
            cell.fill      = piv_hdr_fill
            cell.font      = piv_hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)
            cell.border    = border

        # pivot data rows
        total_row_excel = len(pivot) + 1   # +1 for header row
        for row_idx, row in enumerate(wp.iter_rows(min_row=2), start=2):
            is_total  = (row_idx == total_row_excel + 1)
            base_fill = total_fill if is_total else (piv_alt if row_idx % 2 == 0 else piv_wht)
            for cell in row:
                cell.fill      = base_fill
                cell.font      = total_font if is_total else Font(name="Calibri", size=10)
                cell.alignment = Alignment(
                    horizontal="left" if cell.column == 1 else "center",
                    vertical="center"
                )
                cell.border    = border

        # pivot column widths
        for col_cells in wp.columns:
            max_len = max(
                (len(str(c.value)) if c.value is not None else 0 for c in col_cells),
                default=10,
            )
            wp.column_dimensions[
                get_column_letter(col_cells[0].column)
            ].width = min(max_len + 4, 35)

        wp.row_dimensions[1].height = 30
        wp.freeze_panes = "A2"

        # ── Sheet 3 : Charts ──────────────────────────────────────────────────
        from openpyxl.chart.label import DataLabelList
        from openpyxl.chart.axis import ChartLines

        wb = writer.book
        wc = wb.create_sheet("Charts")

        # ── Shared table styles ───────────────────────────────────────────────
        tbl_hdr_fill  = PatternFill("solid", fgColor="1F3864")
        tbl_hdr_font  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        tbl_body_font = Font(name="Calibri", size=10)
        tbl_alt       = PatternFill("solid", fgColor="EEF2FF")
        tbl_wht       = PatternFill("solid", fgColor="FFFFFF")
        chart_border  = Border(left=thin, right=thin, top=thin, bottom=thin)

        # ── Vivid distinct palettes ───────────────────────────────────────────
        # Region: rich jewel tones — one per region bar
        REGION_COLORS = ["2196F3","F44336","8BC34A","9C27B0","00BCD4","FF9800"]
        # Order Type: bold contrast set
        OT_COLORS     = ["E63946","2196F3","2A9D8F","E9C46A","9C27B0","FF9800"]
        # Order Status: warm-to-cool ramp
        OS_COLORS     = ["2196F3","F44336","8BC34A","9C27B0","FF9800","00BCD4"]

        def apply_datalabels(series):
            """Bold data labels showing value."""
            from openpyxl.chart.label import DataLabelList
            series.dLbls = DataLabelList()
            series.dLbls.showVal        = True
            series.dLbls.showLegendKey  = False
            series.dLbls.showCatName    = False
            series.dLbls.showSerName    = False
            series.dLbls.showPercent    = False

        def style_chart(ch, title, x_title, y_title, is_horizontal=False):
            """Full polish: title, axes, gridlines, legend, dark style."""
            ch.title   = title
            ch.style   = 26         # dark glossy (matches screenshot)
            ch.width   = 22
            ch.height  = 15

            # Axis titles & formatting
            if is_horizontal:
                ch.x_axis.title = x_title   # value axis (count)
                ch.y_axis.title = y_title   # category axis (names)
                ch.x_axis.majorGridlines = ChartLines()
                ch.y_axis.majorGridlines = None
            else:
                ch.x_axis.title = x_title   # category axis (region)
                ch.y_axis.title = y_title   # value axis (count/lines)
                ch.y_axis.majorGridlines = ChartLines()
                ch.x_axis.majorGridlines = None

            ch.x_axis.numFmt = "General"
            ch.y_axis.numFmt = "General"

            # Legend at bottom
            from openpyxl.chart.legend import Legend
            ch.legend          = Legend()
            ch.legend.position = "b"

        def write_table(ws_ref, start_row, start_col, headers, rows, title):
            """Write a styled mini summary table."""
            title_cell = ws_ref.cell(row=start_row, column=start_col, value=title)
            title_cell.font      = Font(bold=True, color="1F3864", name="Calibri", size=11)
            title_cell.alignment = Alignment(horizontal="left", vertical="center")
            ws_ref.merge_cells(
                start_row=start_row, start_column=start_col,
                end_row=start_row,   end_column=start_col + len(headers) - 1
            )
            ws_ref.row_dimensions[start_row].height = 20

            hdr_row = start_row + 1
            for ci, h in enumerate(headers, start=start_col):
                c = ws_ref.cell(row=hdr_row, column=ci, value=h)
                c.fill      = tbl_hdr_fill
                c.font      = tbl_hdr_font
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border    = chart_border
            ws_ref.row_dimensions[hdr_row].height = 18

            for ri, row_data in enumerate(rows):
                dr   = hdr_row + 1 + ri
                fill = tbl_alt if ri % 2 == 0 else tbl_wht
                for ci, val in enumerate(row_data, start=start_col):
                    c = ws_ref.cell(row=dr, column=ci, value=val)
                    c.fill      = fill
                    c.font      = tbl_body_font
                    c.alignment = Alignment(
                        horizontal="left" if ci == start_col else "center",
                        vertical="center"
                    )
                    c.border    = chart_border
                ws_ref.row_dimensions[dr].height = 16

            end_row = hdr_row + len(rows)
            end_col = start_col + len(headers) - 1

            for ci, h in enumerate(headers, start=start_col):
                col_letter = get_column_letter(ci)
                col_vals   = [h] + [str(r[ci - start_col]) for r in rows]
                max_w      = max(len(str(v)) for v in col_vals)
                if ws_ref.column_dimensions[col_letter].width < max_w + 4:
                    ws_ref.column_dimensions[col_letter].width = min(max_w + 4, 30)

            return hdr_row, end_row, start_col, end_col

        # ──────────────────────────────────────────────────────────────────────
        # TABLE 1 — Region: Total Lines + TIM Count
        # ──────────────────────────────────────────────────────────────────────
        region_data = build_region_data(df)
        reg_rows    = [
            (r["Region"], int(r["Total Lines"]), int(r["TIM Count"]))
            for _, r in region_data.iterrows()
        ]
        reg_hdr_row, reg_end_row, reg_sc, reg_ec = write_table(
            wc, start_row=1, start_col=1,
            headers=["Region", "Total Lines", "TIM Count"],
            rows=reg_rows,
            title="Region Summary"
        )

        cats1  = Reference(wc, min_col=reg_sc,
                           min_row=reg_hdr_row + 1, max_row=reg_end_row)

        # Chart 1 — Total Lines per Region
        from openpyxl.chart.series import DataPoint as DP
        chart1 = BarChart()
        chart1.type     = "col"
        chart1.grouping = "clustered"
        chart1.width    = 22
        chart1.height   = 15
        style_chart(chart1,
                    title="Region — Total Lines (MG + Oversold)",
                    x_title="Region", y_title="Total Lines",
                    is_horizontal=False)

        tl_ref = Reference(wc, min_col=reg_sc + 1,
                           min_row=reg_hdr_row, max_row=reg_end_row)
        chart1.add_data(tl_ref, titles_from_data=True)
        chart1.set_categories(cats1)
        for si, color in enumerate(REGION_COLORS[:len(reg_rows)]):
            pt = DP(idx=si)
            pt.graphicalProperties.solidFill = color
            pt.graphicalProperties.line.solidFill = color
            chart1.series[0].dPt.append(pt)
        apply_datalabels(chart1.series[0])
        wc.add_chart(chart1, "F1")

        # Chart 2 — TIM Count per Region
        chart2 = BarChart()
        chart2.type     = "col"
        chart2.grouping = "clustered"
        chart2.width    = 22
        chart2.height   = 15
        style_chart(chart2,
                    title="Region — TIM Count",
                    x_title="Region", y_title="TIM Count",
                    is_horizontal=False)

        tim_ref = Reference(wc, min_col=reg_sc + 2,
                            min_row=reg_hdr_row, max_row=reg_end_row)
        chart2.add_data(tim_ref, titles_from_data=True)
        chart2.set_categories(cats1)
        for si, color in enumerate(REGION_COLORS[:len(reg_rows)]):
            pt = DP(idx=si)
            pt.graphicalProperties.solidFill = color
            pt.graphicalProperties.line.solidFill = color
            chart2.series[0].dPt.append(pt)
        apply_datalabels(chart2.series[0])
        wc.add_chart(chart2, "F26")

        # ──────────────────────────────────────────────────────────────────────
        # TABLE 2 — Order Type counts
        # ──────────────────────────────────────────────────────────────────────
        ot_data  = build_order_type_data(df)
        ot_start = reg_end_row + 3
        ot_rows  = [(r["Order Type"], int(r["Count"])) for _, r in ot_data.iterrows()]
        ot_hdr_row, ot_end_row, ot_sc, ot_ec = write_table(
            wc, start_row=ot_start, start_col=1,
            headers=["Order Type", "Count"],
            rows=ot_rows,
            title="Order Type Summary"
        )

        # Chart 3 — Order Type horizontal bar
        chart3 = BarChart()
        chart3.type     = "bar"
        chart3.grouping = "clustered"
        chart3.width    = 22
        chart3.height   = 15
        style_chart(chart3,
                    title="Order Type — Count Distribution",
                    x_title="Count", y_title="Order Type",
                    is_horizontal=True)

        ot_cats = Reference(wc, min_col=ot_sc,
                            min_row=ot_hdr_row + 1, max_row=ot_end_row)
        ot_vals = Reference(wc, min_col=ot_sc + 1,
                            min_row=ot_hdr_row, max_row=ot_end_row)
        chart3.add_data(ot_vals, titles_from_data=True)
        chart3.set_categories(ot_cats)
        for si, color in enumerate(OT_COLORS[:len(ot_rows)]):
            pt = DP(idx=si)
            pt.graphicalProperties.solidFill = color
            pt.graphicalProperties.line.solidFill = color
            chart3.series[0].dPt.append(pt)
        apply_datalabels(chart3.series[0])
        wc.add_chart(chart3, "F51")

        # ──────────────────────────────────────────────────────────────────────
        # TABLE 3 — Order Status counts
        # ──────────────────────────────────────────────────────────────────────
        os_data  = build_order_status_data(df)
        os_start = ot_end_row + 3
        os_rows  = [(r["Order Staus"], int(r["Count"])) for _, r in os_data.iterrows()]
        os_hdr_row, os_end_row, os_sc, os_ec = write_table(
            wc, start_row=os_start, start_col=1,
            headers=["Order Status", "Count"],
            rows=os_rows,
            title="Order Status Summary"
        )

        # Chart 4 — Order Status horizontal bar
        chart4 = BarChart()
        chart4.type     = "bar"
        chart4.grouping = "clustered"
        chart4.width    = 22
        chart4.height   = 15
        style_chart(chart4,
                    title="Order Status — Count Distribution",
                    x_title="Count", y_title="Order Status",
                    is_horizontal=True)

        os_cats = Reference(wc, min_col=os_sc,
                            min_row=os_hdr_row + 1, max_row=os_end_row)
        os_vals = Reference(wc, min_col=os_sc + 1,
                            min_row=os_hdr_row, max_row=os_end_row)
        chart4.add_data(os_vals, titles_from_data=True)
        chart4.set_categories(os_cats)
        for si, color in enumerate(OS_COLORS[:len(os_rows)]):
            pt = DP(idx=si)
            pt.graphicalProperties.solidFill = color
            pt.graphicalProperties.line.solidFill = color
            chart4.series[0].dPt.append(pt)
        apply_datalabels(chart4.series[0])
        wc.add_chart(chart4, "F76")

        # ── Charts sheet cosmetics ────────────────────────────────────────────
        wc.sheet_view.showGridLines = False
        wc.sheet_properties.tabColor = "1F3864"

    return output.getvalue()


# ══════════════════════════════════════════════════════════════════════════════
#  STREAMLIT UI
# ══════════════════════════════════════════════════════════════════════════════

st.set_page_config(
    page_title="MG Daily Tracker Cleaner",
    page_icon="📊",
    layout="centered",
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=Space+Mono:wght@400;700&display=swap');

html, body, * { font-family: 'Inter', sans-serif; }

[data-testid="stAppViewContainer"] { background: #0B0F1A; }
[data-testid="stHeader"]           { background: transparent; }
section[data-testid="stSidebar"]   { background: #0D1120; }

/* hero */
.hero { margin-bottom: 2rem; }
.hero-badge {
    display: inline-block;
    background: #151C30;
    border: 1px solid #252E4A;
    border-radius: 100px;
    padding: 3px 14px;
    font-size: 0.7rem;
    font-family: 'Space Mono', monospace;
    color: #5B7BFF;
    letter-spacing: 1.5px;
    margin-bottom: 14px;
}
.hero-title {
    font-family: 'Space Mono', monospace;
    font-size: 1.85rem;
    font-weight: 700;
    color: #E4EAFF;
    line-height: 1.2;
    margin-bottom: 6px;
}
.hero-sub {
    font-size: 0.92rem;
    color: #5A6A8A;
    margin: 0;
}

/* cards */
.card {
    background: #111827;
    border: 1px solid #1C2640;
    border-radius: 14px;
    padding: 24px 28px;
    margin-bottom: 18px;
}
.card-accent { border-left: 4px solid #5B7BFF; }
.step-num {
    font-size: 0.65rem;
    font-family: 'Space Mono', monospace;
    color: #5B7BFF;
    letter-spacing: 2px;
    text-transform: uppercase;
    margin-bottom: 6px;
}
.card-title {
    font-size: 1rem;
    font-weight: 600;
    color: #C5D0F0;
    margin-bottom: 4px;
}
.card-desc {
    font-size: 0.82rem;
    color: #4A5878;
    line-height: 1.55;
}

/* rules grid */
.rules-wrap { display: grid; grid-template-columns: 1fr 1fr; gap: 8px; margin-top: 10px; }
.rule-pill {
    background: #131A2E;
    border: 1px solid #1E2A45;
    border-radius: 8px;
    padding: 9px 12px;
    font-size: 0.76rem;
    color: #6A7FA8;
    line-height: 1.45;
}
.rule-pill b { color: #B8C8F0; display: block; margin-bottom: 2px; }

/* file info */
.file-card {
    background: #0E1624;
    border: 1px solid #1A2640;
    border-radius: 10px;
    padding: 16px 20px;
    margin: 12px 0;
    display: flex;
    align-items: center;
    gap: 12px;
}
.file-icon { font-size: 1.5rem; }
.file-meta { flex: 1; }
.file-name { font-weight: 600; color: #C5D0F0; font-size: 0.9rem; }
.file-detail { color: #3D5070; font-size: 0.78rem; margin-top: 2px; }

/* success */
.success-card {
    background: linear-gradient(135deg, #091A10 0%, #0C2015 100%);
    border: 1px solid #164D28;
    border-radius: 12px;
    padding: 18px 22px;
    color: #34D399;
    font-family: 'Space Mono', monospace;
    font-size: 0.88rem;
    margin: 14px 0;
}
.warn-note { color: #FBBF24; font-size: 0.8rem; margin-top: 6px; font-family: 'Inter', sans-serif; }

/* divider */
.hdiv { border: none; border-top: 1px solid #151E30; margin: 22px 0; }
</style>
""", unsafe_allow_html=True)

# ── Header ────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="hero">
  <div class="hero-badge">📊 Make Good · INTERNAL TOOL</div>
  <div class="hero-title">Daily Tracker Cleaner</div>
  <p class="hero-sub">Upload the raw messy tracker → get a clean, formatted Excel file instantly.</p>
</div>
""", unsafe_allow_html=True)

# ── Computed date info ────────────────────────────────────────────────────────
date_str, month_str, current_year = get_date_info()
output_filename   = f"MG Daily Tracker ({date_str}).xlsx"
tracker_tab_name  = f"MG Daily Tracker {date_str}"

# ── Step 1 : Upload ───────────────────────────────────────────────────────────
st.markdown("""
<div class="card card-accent">
  <div class="step-num">Step 01</div>
  <div class="card-title">Upload Your Messy Tracker</div>
  <div class="card-desc">Select the unclean Excel file (.xlsx or .xls) from your computer.</div>
</div>
""", unsafe_allow_html=True)

uploaded_file = st.file_uploader("", type=["xlsx", "xls"], label_visibility="collapsed")

# ── Rules expander ────────────────────────────────────────────────────────────
with st.expander("📋 Cleaning rules reference"):
    st.markdown("""
    <div class="rules-wrap">
      <div class="rule-pill"><b>Date</b>Always set to current date − 1 (DD-Mon-YYYY)</div>
      <div class="rule-pill"><b>Month</b>Current month; if today = 1st → previous month</div>
      <div class="rule-pill"><b>Name</b>Fuzzy-matched to list · empty → "-" · unknown → 🔴 red</div>
      <div class="rule-pill"><b>Region</b>Fuzzy-matched to list · empty → "-" · unknown → 🔴 red</div>
      <div class="rule-pill"><b>TIM#</b>Strip spaces · digits only · empty → "-" · text → 🔴 red</div>
      <div class="rule-pill"><b>CM</b>Left as-is · empty → "-"</div>
      <div class="rule-pill"><b>Parameters</b>Left as-is · empty → "-"</div>
      <div class="rule-pill"><b>MG / Oversold Lines</b>Blank or invalid → 0</div>
      <div class="rule-pill"><b>Total MGL $</b>Currency format · invalid → $0.00</div>
      <div class="rule-pill"><b>Order Type</b>Keyword map + fuzzy · empty → "-" · unknown → 🔴 red</div>
      <div class="rule-pill"><b>Order Status</b>Fuzzy-matched · empty → "-" · unknown → 🔴 red</div>
      <div class="rule-pill"><b>Reason</b>Left as-is · empty → "-"</div>
      <div class="rule-pill"><b>Quality Met</b>Always → Yes</div>
      <div class="rule-pill"><b>Error Type / Comments</b>Always → -</div>
      <div class="rule-pill"><b>TAT Miss</b>Always → No</div>
      <div class="rule-pill"><b>Feedback Delivered by</b>Always → -</div>
      <div class="rule-pill"><b>Year</b>Always current year</div>
    </div>
    """, unsafe_allow_html=True)

# ── Settings expander ─────────────────────────────────────────────────────────
with st.expander("⚙️ Customise lookup lists & sensitivity"):
    c1, c2 = st.columns(2)
    with c1:
        names_txt   = st.text_area("✅ Valid Names (one per line)",
                                    value="\n".join(DEFAULT_NAMES), height=200)
        regions_txt = st.text_area("✅ Valid Regions (one per line)",
                                    value="\n".join(DEFAULT_REGIONS), height=140)
    with c2:
        otypes_txt  = st.text_area("✅ Valid Order Types (one per line)",
                                    value="\n".join(DEFAULT_ORDER_TYPES), height=160)
        ostatus_txt = st.text_area("✅ Valid Order Statuses (one per line)",
                                    value="\n".join(DEFAULT_ORDER_STATUSES), height=180)

    sensitivity = st.slider(
        "Fuzzy match sensitivity (%)",
        min_value=50, max_value=100, value=FUZZY_THRESHOLD,
        help="Higher = stricter. Lower = more aggressive corrections."
    )

# ── Parse customised lists ────────────────────────────────────────────────────
def parse_list(txt):
    return [l.strip() for l in txt.split("\n") if l.strip()]

valid_names    = parse_list(names_txt)
valid_regions  = parse_list(regions_txt)
valid_otypes   = parse_list(otypes_txt)
valid_ostatus  = parse_list(ostatus_txt)

# ── Processing ────────────────────────────────────────────────────────────────
if uploaded_file:
    try:
        raw_df = pd.read_excel(uploaded_file)

        st.markdown(f"""
        <div class="file-card">
          <div class="file-icon">📂</div>
          <div class="file-meta">
            <div class="file-name">{uploaded_file.name}</div>
            <div class="file-detail">{len(raw_df):,} rows · {len(raw_df.columns)} columns detected</div>
          </div>
        </div>
        """, unsafe_allow_html=True)

        with st.expander("👁 Preview raw data (first 5 rows)"):
            st.dataframe(raw_df.head(), use_container_width=True)

        st.markdown('<hr class="hdiv">', unsafe_allow_html=True)

        st.markdown("""
        <div class="card card-accent">
          <div class="step-num">Step 02</div>
          <div class="card-title">Run Cleaning Process</div>
          <div class="card-desc">Click the button to apply all rules and generate the cleaned file.</div>
        </div>
        """, unsafe_allow_html=True)

        if st.button("🧹  Clean Tracker", use_container_width=True, type="primary"):
            with st.spinner("Applying cleaning rules…"):
                cleaned_df, red_cells = clean_tracker(
                    raw_df.copy(),
                    valid_names=valid_names,
                    valid_regions=valid_regions,
                    valid_order_types=valid_otypes,
                    valid_order_statuses=valid_ostatus,
                    threshold=sensitivity,
                )
                pivot_df    = build_pivot(cleaned_df)
                excel_bytes = to_excel_bytes(
                    cleaned_df, red_cells, pivot_df, tracker_tab_name
                )

            red_count = len(red_cells)
            warn_html = (
                f'<div class="warn-note">⚠️ {red_count} cell(s) flagged red — '
                f'review them manually in the downloaded file.</div>'
                if red_count else ""
            )

            st.markdown(f"""
            <div class="success-card">
              ✅ Cleaning complete — {len(cleaned_df):,} rows processed
              {warn_html}
            </div>
            """, unsafe_allow_html=True)

            with st.expander("👁 Preview cleaned data (first 5 rows)"):
                st.dataframe(cleaned_df.head(), use_container_width=True)

            with st.expander("👁 Preview pivot table"):
                st.dataframe(pivot_df, use_container_width=True, hide_index=True)

            st.markdown('<hr class="hdiv">', unsafe_allow_html=True)

            st.markdown("""
            <div class="card card-accent">
              <div class="step-num">Step 03</div>
              <div class="card-title">Download Cleaned Tracker</div>
              <div class="card-desc">
                Three sheets inside: <b>Cleaned Data</b> · <b>Pivot</b> · <b>Charts</b>.<br>
                Red cells = values the cleaner could not confidently correct — fill those manually.
              </div>
            </div>
            """, unsafe_allow_html=True)

            st.download_button(
                label=f"⬇️   Download  {output_filename}",
                data=excel_bytes,
                file_name=output_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

    except Exception as e:
        st.error(f"❌ Could not read the file: {e}")
        st.info("Please make sure you upload a valid Excel file (.xlsx or .xls).")

else:
    st.markdown("""
    <div style="text-align:center;padding:52px 0;color:#1E2D48;font-size:0.88rem;">
        ↑ Upload a file above to begin
    </div>
    """, unsafe_allow_html=True)

# ── Footer ────────────────────────────────────────────────────────────────────
st.markdown("""
<div style="margin-top:56px;text-align:center;font-size:0.68rem;
            color:#1A2438;font-family:'Space Mono',monospace;letter-spacing:1px;">
    Make Good INTERNAL · DAILY TRACKER AUTOMATION
</div>
""", unsafe_allow_html=True)
